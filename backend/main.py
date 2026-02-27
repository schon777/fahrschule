import base64
import hashlib
import json
import os
import sqlite3
import uuid
import io
import ipaddress
import math
import re
import time
import random
import zipfile
import urllib.parse
import urllib.request
from html.parser import HTMLParser
from datetime import datetime, timedelta, timezone

import jwt
from fastapi import FastAPI, HTTPException, Depends, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware

BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, "app.db")
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


CONFIG = load_config()

OFFLINE_MODE = bool(CONFIG.get("offline_mode", False))

OLLAMA_BASE = CONFIG.get("ollama_base", "http://127.0.0.1:11434")
AI_MODELS = {
    "chat": CONFIG.get("ai_model_chat", "llama3.1:8b"),
    "embed": CONFIG.get("ai_model_embed", "nomic-embed-text"),
    "quiz": CONFIG.get("ai_model_quiz", "mistral:7b"),
    "grade": CONFIG.get("ai_model_grade", "qwen2.5:7b"),
}
AI_ALLOWLIST = CONFIG.get(
    "ai_allowlist",
    ["wikipedia.org", "ihk.de", "docs.python.org", "developer.mozilla.org"],
)
AI_BROWSE_MAX_PER_DAY = int(CONFIG.get("ai_browse_per_day", 30))

_rate_limits = {}


class HTMLTextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self._chunks = []

    def handle_data(self, data):
        if data and data.strip():
            self._chunks.append(data.strip())

    def get_text(self):
        return " ".join(self._chunks)


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_table_columns(conn, table):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return {row["name"] for row in cur.fetchall()}


def ensure_column(conn, table, column, col_type, default=None):
    columns = get_table_columns(conn, table)
    if column in columns:
        return
    cur = conn.cursor()
    cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
    if default is not None:
        cur.execute(f"UPDATE {table} SET {column} = ? WHERE {column} IS NULL", (default,))
    conn.commit()


def ensure_topic_paths(conn):
    cur = conn.cursor()
    cur.execute("SELECT id, parent_id, path, depth FROM topics")
    rows = [dict(row) for row in cur.fetchall()]
    if not rows:
        return
    updates = []
    for row in rows:
        if row.get("path") and row.get("depth") is not None:
            continue
        updates.append((row["id"],))
    if not updates:
        return
    cur.execute("SELECT id FROM topics")
    existing = {r["id"] for r in cur.fetchall()}
    for (topic_id,) in updates:
        if topic_id not in existing:
            continue
        cur.execute(
            "UPDATE topics SET parent_id = NULL, path = ?, depth = 0 WHERE id = ?",
            (topic_id, topic_id),
        )
    conn.commit()


def build_topic_map(topics_input, errors):
    topic_map = {}
    for t in topics_input:
        slug = t.get("slug") or t.get("id") or t.get("topic_id")
        if slug:
            parent_id = t.get("parent_topic_id") or t.get("parent_id")
            topic_map[slug] = {
                "id": slug,
                "name": t.get("title") or t.get("name") or slug_to_title(slug),
                "topic_area": t.get("topic_area", ""),
                "parent_id": parent_id,
                "path": t.get("path"),
                "depth": t.get("depth"),
            }
        else:
            errors.append("Topic missing slug/id.")

    # Build paths/depths for topic tree.
    visiting = set()
    visited = set()

    def compute_tree(slug):
        if slug in visited:
            return
        if slug in visiting:
            errors.append(f"Topic cycle detected at {slug}.")
            return
        visiting.add(slug)
        node = topic_map.get(slug, {})
        parent_id = node.get("parent_id")
        if parent_id and parent_id not in topic_map:
            topic_map[parent_id] = {
                "id": parent_id,
                "name": slug_to_title(parent_id),
                "topic_area": "",
                "parent_id": None,
                "path": parent_id,
                "depth": 0,
            }
        if parent_id:
            compute_tree(parent_id)
            parent = topic_map[parent_id]
            if not node.get("path"):
                node["path"] = f"{parent['path']}/{slug}"
            if node.get("depth") is None:
                node["depth"] = (parent.get("depth") or 0) + 1
        else:
            node["path"] = node.get("path") or slug
            node["depth"] = node.get("depth") if node.get("depth") is not None else 0
        topic_map[slug] = node
        visiting.remove(slug)
        visited.add(slug)

    for slug in list(topic_map.keys()):
        compute_tree(slug)

    return topic_map


def get_pack_id(pack):
    meta = pack.get("meta", {}) if isinstance(pack, dict) else {}
    pack_id = meta.get("pack_id") or meta.get("id")
    if pack_id:
        return str(pack_id)
    seed = f"{meta.get('title','')}-{meta.get('created_at','')}"
    digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12]
    return f"pack_{digest}"


def upsert_pack(conn, pack_id, pack):
    meta = pack.get("meta", {}) if isinstance(pack, dict) else {}
    settings = pack.get("settings", {}) if isinstance(pack, dict) else {}
    methods = pack.get("methods", []) if isinstance(pack, dict) else []
    assets = pack.get("assets", []) if isinstance(pack, dict) else []
    schema = pack.get("schema", "")
    created_at = meta.get("created_at") or datetime.now(timezone.utc).isoformat()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM packs WHERE id = ?", (pack_id,))
    exists = cur.fetchone() is not None
    payload = (
        pack_id,
        schema,
        json.dumps(meta),
        json.dumps(settings),
        json.dumps(methods),
        json.dumps(assets),
        created_at,
    )
    if exists:
        cur.execute(
            """
            UPDATE packs
            SET schema = ?, meta = ?, settings = ?, methods = ?, assets = ?, created_at = ?
            WHERE id = ?
            """,
            (schema, json.dumps(meta), json.dumps(settings), json.dumps(methods), json.dumps(assets), created_at, pack_id),
        )
    else:
        cur.execute(
            """
            INSERT INTO packs (id, schema, meta, settings, methods, assets, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            payload,
        )


def safe_list(value):
    return value if isinstance(value, list) else []


def normalize_quiztab_question(input_q, topic_map, errors, pack_id):
    if not input_q or "id" not in input_q or "method_id" not in input_q:
        errors.append("Quiztab question missing id or method_id.")
        return None
    method_id_raw = input_q.get("method_id")
    type_map = {"guessword": "guess", "explainterm": "explain"}
    method_id = type_map.get(method_id_raw, method_id_raw)
    topic_slug = input_q.get("topic_slug")
    if not topic_slug:
        errors.append(f"Question {input_q['id']} missing topic_slug.")
        return None
    if topic_slug not in topic_map:
        topic_map[topic_slug] = {
            "id": topic_slug,
            "name": slug_to_title(topic_slug),
            "topic_area": "",
        }
    prompt = input_q.get("prompt")
    if not prompt:
        errors.append(f"Question {input_q['id']} missing prompt.")
        return None

    payload = input_q.get("payload") if isinstance(input_q.get("payload"), dict) else {}
    support = input_q.get("support") if isinstance(input_q.get("support"), dict) else {}
    solution = input_q.get("solution") if isinstance(input_q.get("solution"), dict) else {}

    base = {
        "id": input_q["id"],
        "topicId": topic_slug,
        "type": method_id,
        "method_id": method_id,
        "prompt": prompt,
        "payload": payload,
        "support": support,
        "solution": solution,
        "difficulty": input_q.get("difficulty", ""),
        "variants": safe_list(input_q.get("variants")),
        "randomization": input_q.get("randomization"),
        "source_ref": input_q.get("source_ref", "internal:import"),
        "tags": safe_list(input_q.get("tags")),
        "pack_id": pack_id,
    }

    # Legacy compatibility for existing renderer/logic.
    if method_id == "single":
        options = payload.get("options") if "options" in payload else input_q.get("options")
        if "correct" in payload:
            correct = payload.get("correct")
        elif "correctIndexes" in payload:
            correct = payload.get("correctIndexes")
        else:
            correct = input_q.get("correct")
        if isinstance(options, list):
            base["options"] = options
        if isinstance(correct, list) and len(correct) > 0:
            try:
                base["correctIndex"] = int(correct[0])
            except (TypeError, ValueError):
                pass
        elif isinstance(correct, int):
            base["correctIndex"] = int(correct)
        elif isinstance(correct, str) and correct.strip().isdigit():
            base["correctIndex"] = int(correct.strip())
    if method_id == "multi":
        options = payload.get("options") if "options" in payload else input_q.get("options")
        if "correct" in payload:
            correct = payload.get("correct")
        elif "correctIndexes" in payload:
            correct = payload.get("correctIndexes")
        else:
            correct = input_q.get("correct")
        if isinstance(options, list):
            base["options"] = options
        if isinstance(correct, list):
            parsed = []
            for item in correct:
                try:
                    parsed.append(int(item))
                except (TypeError, ValueError):
                    continue
            if parsed:
                base["correctIndexes"] = parsed
        elif isinstance(correct, int):
            base["correctIndexes"] = [int(correct)]
        elif isinstance(correct, str) and correct.strip().isdigit():
            base["correctIndexes"] = [int(correct.strip())]
    if method_id == "truefalse":
        val = payload.get("correct") if "correct" in payload else input_q.get("correct")
        if isinstance(val, bool):
            base["correctBoolean"] = val
    if method_id == "fillblank":
        blanks = payload.get("blanks") or input_q.get("answers") or input_q.get("expectedAnswers")
        if isinstance(blanks, list):
            base["answers"] = blanks
    if method_id == "matching":
        pairs = payload.get("pairs")
        left = payload.get("left")
        right = payload.get("right")
        if isinstance(pairs, list) and len(pairs) > 0 and isinstance(pairs[0], dict):
            base["pairs"] = pairs
        elif isinstance(left, list) and isinstance(right, list) and isinstance(pairs, list):
            mapped = []
            for pair in pairs:
                try:
                    l_idx, r_idx = pair
                    mapped.append({"left": left[l_idx], "right": right[r_idx]})
                except Exception:
                    continue
            if mapped:
                base["pairs"] = mapped
    if method_id == "ordering":
        items = payload.get("items")
        correct_order = payload.get("correct_order")
        if isinstance(items, list) and isinstance(correct_order, list):
            base["items"] = items
            base["correct_order"] = correct_order
    if method_id == "guess":
        answers = payload.get("answers") or input_q.get("answers")
        if isinstance(answers, list):
            base["expectedAnswers"] = answers
    if method_id == "explain":
        expected = payload.get("expectedAnswer") or solution.get("final") or input_q.get("expectedAnswer", "")
        if isinstance(expected, str):
            base["expectedAnswer"] = expected
    if method_id == "exam":
        expected = payload.get("expectedAnswer") or solution.get("final") or input_q.get("expectedAnswer", "")
        if isinstance(expected, str):
            base["expectedAnswer"] = expected

    return base


def normalize_number(text):
    if isinstance(text, (int, float)):
        return float(text)
    if not isinstance(text, str):
        return None
    cleaned = text.strip().replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


UNIT_TABLE = {
    "a": ("a", 1.0),
    "ma": ("a", 0.001),
    "ua": ("a", 0.000001),
    "w": ("w", 1.0),
    "kw": ("w", 1000.0),
    "va": ("va", 1.0),
    "kva": ("va", 1000.0),
    "v": ("v", 1.0),
    "mv": ("v", 0.001),
    "ohm": ("ohm", 1.0),
}


def parse_value_unit(value, unit):
    raw_unit = unit
    if unit is None and isinstance(value, str):
        parts = value.strip().split()
        if len(parts) == 2:
            value, unit = parts[0], parts[1]
            raw_unit = unit
        else:
            match = re.match(r"^([0-9\.,]+)\s*([a-zA-Z]+)$", value.strip())
            if match:
                value, unit = match.group(1), match.group(2)
                raw_unit = unit
    num = normalize_number(value)
    if num is None:
        return None, None, raw_unit, "Invalid number"
    if not unit:
        return num, None, raw_unit, None
    key = str(unit).strip().lower()
    if key not in UNIT_TABLE:
        return num, None, raw_unit, "Unknown unit"
    base_unit, factor = UNIT_TABLE[key]
    return num * factor, base_unit, raw_unit, None


def round_value(value, decimals):
    try:
        return round(float(value), int(decimals))
    except (TypeError, ValueError):
        return value


def validate_calc_value_payload(payload):
    expected = payload.get("expected_value")
    unit = payload.get("expected_unit")
    if not isinstance(expected, (int, float)):
        return False, "calc_value requires expected_value"
    if not isinstance(unit, str) or not unit:
        return False, "calc_value requires expected_unit"
    return True, ""


def validate_calc_multi_payload(payload):
    fields = payload.get("fields")
    answers = payload.get("answers")
    if not isinstance(fields, list) or len(fields) == 0:
        return False, "calc_multi requires fields"
    if not isinstance(answers, dict) or len(answers) == 0:
        return False, "calc_multi requires answers"
    return True, ""


def validate_hotspot_payload(payload):
    svg = payload.get("svg")
    hotspots = payload.get("hotspots")
    correct = payload.get("correct")
    if not isinstance(svg, str) or not svg.strip():
        return False, "hotspot_svg requires svg"
    if not isinstance(hotspots, list) or len(hotspots) == 0:
        return False, "hotspot_svg requires hotspots"
    if not isinstance(correct, list):
        return False, "hotspot_svg requires correct"
    return True, ""


def validate_troubleshoot_payload(payload):
    start = payload.get("start")
    nodes = payload.get("nodes")
    success = payload.get("success_node")
    if not start or not isinstance(nodes, list) or len(nodes) == 0 or not success:
        return False, "troubleshoot_flow requires start, nodes, success_node"
    node_ids = {n.get("id") for n in nodes if isinstance(n, dict)}
    if start not in node_ids or success not in node_ids:
        return False, "troubleshoot_flow invalid node references"
    for node in nodes:
        choices = node.get("choices", [])
        for c in choices:
            nxt = c.get("next")
            if nxt and nxt not in node_ids:
                return False, "troubleshoot_flow has invalid choice reference"
    return True, ""

def log_ai(user, action, detail):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO ai_logs (id, username, action, detail, created_at) VALUES (?, ?, ?, ?, ?)",
        (
            f"ailog_{uuid.uuid4().hex}",
            user,
            action,
            json.dumps(detail) if isinstance(detail, (dict, list)) else str(detail),
            datetime.now(timezone.utc).isoformat(),
        ),
    )
    conn.commit()
    conn.close()


def check_rate_limit(user, action):
    if action != "browse":
        return
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    key = (user, action, today)
    current = _rate_limits.get(key, 0)
    if current >= AI_BROWSE_MAX_PER_DAY:
        raise HTTPException(status_code=429, detail="Browse rate limit reached.")
    _rate_limits[key] = current + 1


def is_allowed_url(url):
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in ("http", "https"):
        return False
    host = parsed.hostname or ""
    try:
        ip = ipaddress.ip_address(host)
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved:
            return False
        return True
    except ValueError:
        pass
    host = host.lower()
    return any(host == d or host.endswith(f".{d}") for d in AI_ALLOWLIST)


def fetch_url_text(url, timeout=10):
    if OFFLINE_MODE:
        raise HTTPException(status_code=403, detail="Offline mode enabled")
    if not is_allowed_url(url):
        raise HTTPException(status_code=403, detail="URL not allowed")
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "AP2-AI-Browser/1.0"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read()
        encoding = resp.headers.get_content_charset() or "utf-8"
        html = raw.decode(encoding, errors="ignore")
    parser = HTMLTextExtractor()
    parser.feed(html)
    text = parser.get_text()
    return re.sub(r"\s+", " ", text).strip()


def ollama_request(path, payload, timeout=60):
    url = f"{OLLAMA_BASE}{path}"
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def ollama_generate(prompt, model, system=None, temperature=0.2):
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": temperature},
    }
    if system:
        payload["system"] = system
    result = ollama_request("/api/generate", payload)
    return result.get("response", "").strip()


def ollama_embed(text, model=None):
    payload = {"model": model or AI_MODELS["embed"], "prompt": text}
    result = ollama_request("/api/embeddings", payload)
    return result.get("embedding", [])


def chunk_text(text, max_chars=1200, overlap=200):
    if not text:
        return []
    chunks = []
    start = 0
    length = len(text)
    while start < length:
        end = min(start + max_chars, length)
        chunk = text[start:end]
        chunks.append(chunk)
        if end == length:
            break
        start = max(0, end - overlap)
    return chunks


def cosine_similarity(a, b):
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x * x for x in a))
    norm_b = math.sqrt(sum(x * x for x in b))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)


def retrieve_chunks(query, top_k=5):
    query_emb = ollama_embed(query)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, doc_id, text, embedding FROM chunks")
    scored = []
    for row in cur.fetchall():
        emb = json.loads(row["embedding"]) if row["embedding"] else []
        score = cosine_similarity(query_emb, emb)
        scored.append((score, row["id"], row["doc_id"], row["text"]))
    conn.close()
    scored.sort(reverse=True, key=lambda x: x[0])
    top = scored[:top_k]
    return [
        {"score": s, "chunk_id": cid, "doc_id": did, "text": txt}
        for s, cid, did, txt in top
    ]


def extract_text_from_upload(file: UploadFile):
    filename = file.filename or "upload"
    content_type = file.content_type or "application/octet-stream"
    raw = file.file.read()
    sha256 = hashlib.sha256(raw).hexdigest()
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".txt", ".md"):
        return raw.decode("utf-8", errors="ignore"), content_type, sha256
    if ext in (".html", ".htm"):
        parser = HTMLTextExtractor()
        parser.feed(raw.decode("utf-8", errors="ignore"))
        return parser.get_text(), content_type, sha256
    if ext == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages), content_type, sha256
    if ext == ".docx":
        import docx

        doc = docx.Document(io.BytesIO(raw))
        return "\n".join([p.text for p in doc.paragraphs]), content_type, sha256
    raise HTTPException(status_code=400, detail="Unsupported file type")


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS topics (
            id TEXT PRIMARY KEY,
            name TEXT,
            topic_area TEXT,
            parent_id TEXT,
            path TEXT,
            depth INTEGER
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS questions (
            id TEXT PRIMARY KEY,
            topic_id TEXT,
            type TEXT,
            data TEXT,
            explanation TEXT,
            source_ref TEXT,
            tags TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS attempts (
            id TEXT PRIMARY KEY,
            username TEXT,
            question_id TEXT,
            selected TEXT,
            correct INTEGER,
            graded_by_user INTEGER,
            timestamp TEXT,
            time_ms INTEGER
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS appointments (
            id TEXT PRIMARY KEY,
            username TEXT,
            title TEXT,
            start TEXT,
            end TEXT,
            notes TEXT,
            category TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS documents (
            id TEXT PRIMARY KEY,
            filename TEXT,
            sha256 TEXT,
            content_type TEXT,
            uploaded_at TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS chunks (
            id TEXT PRIMARY KEY,
            doc_id TEXT,
            chunk_index INTEGER,
            text TEXT,
            embedding TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ai_logs (
            id TEXT PRIMARY KEY,
            username TEXT,
            action TEXT,
            detail TEXT,
            created_at TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS packs (
            id TEXT PRIMARY KEY,
            schema TEXT,
            meta TEXT,
            settings TEXT,
            methods TEXT,
            assets TEXT,
            created_at TEXT
        )
        """
    )
    conn.commit()

    # Backfill older schemas.
    ensure_column(conn, "topics", "parent_id", "TEXT")
    ensure_column(conn, "topics", "path", "TEXT")
    ensure_column(conn, "topics", "depth", "INTEGER")
    ensure_column(conn, "questions", "pack_id", "TEXT", default="default_pack")
    ensure_topic_paths(conn)

    conn.close()


def seed_data():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS count FROM questions")
    if cur.fetchone()["count"] > 0:
        conn.close()
        return

    topics = [
        {"id": "netzwerk", "name": "Netzwerk Grundlagen", "topic_area": "netzwerk"},
        {"id": "elektro", "name": "Elektrotechnik", "topic_area": "elektrotechnik"},
        {"id": "it", "name": "IT Systeme", "topic_area": "it-systeme"},
    ]

    questions = [
        {
            "id": "q1",
            "topicId": "netzwerk",
            "type": "multi",
            "prompt": "Welche Aussagen zu IPv4 sind korrekt?",
            "options": [
                "IPv4 nutzt 32 Bit Adressen",
                "IPv4 hat 128 Bit Adressen",
                "Private Netze nutzen z.B. 192.168.0.0/16",
                "IPv4 kann keine Subnetze nutzen",
            ],
            "correctIndexes": [0, 2],
            "explanation": "IPv4 nutzt 32 Bit und private Netze sind definiert.",
            "source_ref": "Library/Netzwerk/Grundlagen/ip-adressierung-v4-v6.md",
        },
        {
            "id": "q2",
            "topicId": "netzwerk",
            "type": "multi",
            "prompt": "Welche Protokolle sind Layer 2?",
            "options": ["Ethernet", "IP", "TCP", "ARP"],
            "correctIndexes": [0, 3],
            "explanation": "Ethernet und ARP arbeiten am Link Layer.",
            "source_ref": "Library/Netzwerk/Grundlagen/osi-tcpip-ethernet.md",
        },
        {
            "id": "q3",
            "topicId": "elektro",
            "type": "multi",
            "prompt": "Welche Aussagen zum Ohmschen Gesetz stimmen?",
            "options": ["U = R * I", "I = U / R", "R = U * I", "R = U / I"],
            "correctIndexes": [0, 1, 3],
            "explanation": "Ohmsches Gesetz: U = R * I.",
            "source_ref": "Library/Elektrotechnik/Grundlagen/spannung-strom-ohm.md",
        },
        {
            "id": "q4",
            "topicId": "elektro",
            "type": "multi",
            "prompt": "Welche Schutzeinrichtungen koennen Fehlerstrom erfassen?",
            "options": ["RCD", "LS Schalter", "FI", "NH Sicherung"],
            "correctIndexes": [0, 2],
            "explanation": "RCD und FI sind Fehlerstromschutz.",
            "source_ref": "Library/Elektrotechnik/Schutztechnik/rcd-fi-fehlerstromschutz.md",
        },
        {
            "id": "q5",
            "topicId": "it",
            "type": "multi",
            "prompt": "Welche Aussagen zur CIA Triad sind korrekt?",
            "options": [
                "C steht fuer Confidentiality",
                "I steht fuer Integrity",
                "A steht fuer Availability",
                "C steht fuer Compliance",
            ],
            "correctIndexes": [0, 1, 2],
            "explanation": "CIA Triad: Confidentiality, Integrity, Availability.",
            "source_ref": "Library/IT-Systeme/Grundlagen/cia-triad-datenschutz-datensicherheit.md",
        },
        {
            "id": "q6",
            "topicId": "it",
            "type": "multi",
            "prompt": "Welche Aussagen zu Backups stimmen?",
            "options": [
                "3-2-1 Regel: 3 Kopien auf 2 Medien, 1 extern",
                "RAID ist immer ein Backup",
                "Backups sollten regelmaessig getestet werden",
                "Snapshots ersetzen alle Backups",
            ],
            "correctIndexes": [0, 2],
            "explanation": "RAID ist kein Backup, Tests sind wichtig.",
            "source_ref": "Library/IT-Systeme/Dienste/backupverfahren-3-2-1-generationenprinzip.md",
        },
        {
            "id": "tf1",
            "topicId": "netzwerk",
            "type": "truefalse",
            "prompt": "IPv4 nutzt 32 Bit Adressen.",
            "correctBoolean": True,
            "explanation": "IPv4 ist 32 Bit.",
            "source_ref": "Library/Netzwerk/Grundlagen/ip-adressierung-v4-v6.md",
        },
        {
            "id": "tf2",
            "topicId": "netzwerk",
            "type": "truefalse",
            "prompt": "VLANs arbeiten auf Layer 3.",
            "correctBoolean": False,
            "explanation": "VLANs sind Layer 2 Segmentierung.",
            "source_ref": "Library/Netzwerk/VLAN/vlan-grundlagen-trunk-tagging.md",
        },
        {
            "id": "tf3",
            "topicId": "elektro",
            "type": "truefalse",
            "prompt": "RCD und FI sind Fehlerstromschutz.",
            "correctBoolean": True,
            "explanation": "RCD/FI sind Fehlerstromschutz.",
            "source_ref": "Library/Elektrotechnik/Schutztechnik/rcd-fi-fehlerstromschutz.md",
        },
        {
            "id": "tf4",
            "topicId": "elektro",
            "type": "truefalse",
            "prompt": "Der Schutzleiter darf geschaltet werden.",
            "correctBoolean": False,
            "explanation": "Der Schutzleiter darf nicht geschaltet werden.",
            "source_ref": "Library/Elektrotechnik/Sicherheit/fuenf-sicherheitsregeln.md",
        },
        {
            "id": "tf5",
            "topicId": "it",
            "type": "truefalse",
            "prompt": "RAID ersetzt ein Backup.",
            "correctBoolean": False,
            "explanation": "RAID ist kein Backup.",
            "source_ref": "Library/IT-Systeme/Dienste/raid-grundlagen.md",
        },
        {
            "id": "m1",
            "topicId": "netzwerk",
            "type": "matching",
            "prompt": "Match protocol to function.",
            "pairs": [
                {"left": "DHCP", "right": "IP Address Assignment"},
                {"left": "DNS", "right": "Name Resolution"},
                {"left": "NAT", "right": "Address Translation"},
            ],
            "explanation": "These are core network services.",
            "source_ref": "Library/Netzwerk/Dienste/dhcp-dns.md",
        },
        {
            "id": "m2",
            "topicId": "netzwerk",
            "type": "matching",
            "prompt": "Match OSI layer to example.",
            "pairs": [
                {"left": "Layer 2", "right": "Ethernet"},
                {"left": "Layer 3", "right": "IP"},
                {"left": "Layer 4", "right": "TCP"},
            ],
            "explanation": "OSI layers have typical protocols.",
            "source_ref": "Library/Netzwerk/Grundlagen/osi-tcpip-ethernet.md",
        },
        {
            "id": "m3",
            "topicId": "elektro",
            "type": "matching",
            "prompt": "Match device to role.",
            "pairs": [
                {"left": "LS", "right": "Overcurrent protection"},
                {"left": "RCD", "right": "Residual current protection"},
                {"left": "SLS", "right": "Selective main protection"},
            ],
            "explanation": "Protection devices have different roles.",
            "source_ref": "Library/Elektrotechnik/Schutztechnik/selektivitaet-sls.md",
        },
        {
            "id": "m4",
            "topicId": "it",
            "type": "matching",
            "prompt": "Match term to meaning.",
            "pairs": [
                {"left": "CIA", "right": "Security goals"},
                {"left": "VM", "right": "Virtual machine"},
                {"left": "NAS", "right": "Network storage"},
            ],
            "explanation": "Common IT terms.",
            "source_ref": "Library/IT-Systeme/Virtualisierung/virtualisierung-hypervisor-grundlagen.md",
        },
        {
            "id": "m5",
            "topicId": "it",
            "type": "matching",
            "prompt": "Match service to type.",
            "pairs": [
                {"left": "SaaS", "right": "Software as a Service"},
                {"left": "PaaS", "right": "Platform as a Service"},
                {"left": "IaaS", "right": "Infrastructure as a Service"},
            ],
            "explanation": "Cloud service models.",
            "source_ref": "Library/IT-Systeme/Dienste/cloud-services-saas-paas-iaas.md",
        },
        {
            "id": "fb1",
            "topicId": "netzwerk",
            "type": "fillblank",
            "prompt": "Fill in the blank: IPv4 has ___ bits.",
            "expectedAnswers": ["32"],
            "explanation": "IPv4 uses 32-bit addresses.",
            "source_ref": "Library/Netzwerk/Grundlagen/ip-adressierung-v4-v6.md",
        },
        {
            "id": "fb2",
            "topicId": "netzwerk",
            "type": "fillblank",
            "prompt": "Fill in the blank: VLAN tags use standard ___ .",
            "expectedAnswers": ["802.1q", "8021q"],
            "explanation": "VLAN tagging uses IEEE 802.1Q.",
            "source_ref": "Library/Netzwerk/VLAN/vlan-grundlagen-trunk-tagging.md",
        },
        {
            "id": "fb3",
            "topicId": "elektro",
            "type": "fillblank",
            "prompt": "Fill in the blank: U = R * ___.",
            "expectedAnswers": ["I", "i"],
            "explanation": "Ohmsches Gesetz.",
            "source_ref": "Library/Elektrotechnik/Grundlagen/spannung-strom-ohm.md",
        },
        {
            "id": "fb4",
            "topicId": "elektro",
            "type": "fillblank",
            "prompt": "Fill in the blank: A FI is a ___ device.",
            "expectedAnswers": ["fehlerstromschutz", "rcd"],
            "explanation": "FI is a residual current device.",
            "source_ref": "Library/Elektrotechnik/Schutztechnik/rcd-fi-fehlerstromschutz.md",
        },
        {
            "id": "fb5",
            "topicId": "it",
            "type": "fillblank",
            "prompt": "Fill in the blank: CIA stands for Confidentiality, Integrity, and ___.",
            "expectedAnswers": ["availability"],
            "explanation": "CIA triad.",
            "source_ref": "Library/IT-Systeme/Grundlagen/cia-triad-datenschutz-datensicherheit.md",
        },
        {
            "id": "s1",
            "topicId": "netzwerk",
            "type": "single",
            "prompt": "Which port is standard for HTTPS?",
            "options": ["80", "443", "53", "22"],
            "correctIndex": 1,
            "explanation": "HTTPS uses port 443.",
            "source_ref": "Library/Netzwerk/Grundlagen/osi-tcpip-ethernet.md",
        },
        {
            "id": "s2",
            "topicId": "elektro",
            "type": "single",
            "prompt": "Which unit is for electrical current?",
            "options": ["Volt", "Ampere", "Ohm", "Watt"],
            "correctIndex": 1,
            "explanation": "Current is measured in Ampere.",
            "source_ref": "Library/Elektrotechnik/Grundlagen/spannung-strom-ohm.md",
        },
        {
            "id": "s3",
            "topicId": "it",
            "type": "single",
            "prompt": "Which is a backup rule?",
            "options": ["3-2-1", "4-4-2", "1-1-1", "2-3-5"],
            "correctIndex": 0,
            "explanation": "3-2-1 is common backup rule.",
            "source_ref": "Library/IT-Systeme/Dienste/backupverfahren-3-2-1-generationenprinzip.md",
        },
        {
            "id": "g1",
            "topicId": "netzwerk",
            "type": "guess",
            "prompt": "Guess the word: Protocol that assigns IP addresses automatically.",
            "expectedAnswers": ["dhcp"],
            "explanation": "DHCP assigns IP addresses.",
            "source_ref": "Library/Netzwerk/Dienste/dhcp-dns.md",
        },
        {
            "id": "g2",
            "topicId": "elektro",
            "type": "guess",
            "prompt": "Guess the word: Device that protects against residual current.",
            "expectedAnswers": ["rcd", "fi"],
            "explanation": "RCD/FI is residual current device.",
            "source_ref": "Library/Elektrotechnik/Schutztechnik/rcd-fi-fehlerstromschutz.md",
        },
        {
            "id": "g3",
            "topicId": "it",
            "type": "guess",
            "prompt": "Guess the word: Security goals trio (C, I, A).",
            "expectedAnswers": ["cia"],
            "explanation": "CIA triad.",
            "source_ref": "Library/IT-Systeme/Grundlagen/cia-triad-datenschutz-datensicherheit.md",
        },
        {
            "id": "e1",
            "topicId": "netzwerk",
            "type": "explain",
            "prompt": "Explain VLANs in one or two sentences.",
            "expectedAnswer": "Segmentiert Layer 2 Netzwerke logisch, trennt Broadcast Domains.",
            "explanation": "VLANs segment networks at Layer 2.",
            "source_ref": "Library/Netzwerk/VLAN/vlan-grundlagen-trunk-tagging.md",
        },
        {
            "id": "e2",
            "topicId": "elektro",
            "type": "explain",
            "prompt": "Explain the purpose of protective earth (PE).",
            "expectedAnswer": "Sichert gegen Fehlerstrom und leitet ihn ab.",
            "explanation": "PE protects against fault current.",
            "source_ref": "Library/Elektrotechnik/Sicherheit/schutz-gegen-elektrischen-schlag.md",
        },
        {
            "id": "e3",
            "topicId": "it",
            "type": "explain",
            "prompt": "Explain the 3-2-1 backup rule.",
            "expectedAnswer": "3 copies, 2 media, 1 offsite.",
            "explanation": "3-2-1 is a backup best practice.",
            "source_ref": "Library/IT-Systeme/Dienste/backupverfahren-3-2-1-generationenprinzip.md",
        },
        {
            "id": "x1",
            "topicId": "netzwerk",
            "type": "exam",
            "prompt": "Exam scenario: Design a small VLAN plan for 3 departments with 60 clients. Outline VLAN IDs and subnets.",
            "expectedAnswer": "Provide VLAN IDs and /26 or /25 subnets per department.",
            "explanation": "Keep subnets per department and document VLAN IDs.",
            "source_ref": "Library/Netzwerk/Methoden/checkliste-vlan-plan.md",
        },
        {
            "id": "x2",
            "topicId": "elektro",
            "type": "exam",
            "prompt": "Exam scenario: Describe steps for safe work on electrical equipment.",
            "expectedAnswer": "Five safety rules in order.",
            "explanation": "Use five safety rules.",
            "source_ref": "Library/Elektrotechnik/Sicherheit/fuenf-sicherheitsregeln.md",
        },
        {
            "id": "x3",
            "topicId": "it",
            "type": "exam",
            "prompt": "Exam scenario: Plan a backup strategy for a small office.",
            "expectedAnswer": "Define 3-2-1, schedule, and test plan.",
            "explanation": "Use 3-2-1 and test restores.",
            "source_ref": "Library/IT-Systeme/Dienste/backupverfahren-3-2-1-generationenprinzip.md",
        },
        {
            "id": "o1",
            "topicId": "netzwerk",
            "type": "ordering",
            "prompt": "Order the steps for basic network troubleshooting.",
            "items": ["Check link", "Check IP config", "Ping gateway"],
            "correct_order": [0, 1, 2],
            "explanation": "Start physical, then config, then reachability.",
            "source_ref": "Library/Netzwerk/Methoden/troubleshooting-dhcp.md",
        },
        {
            "id": "o2",
            "topicId": "elektro",
            "type": "ordering",
            "prompt": "Order the five safety rules (simplified).",
            "items": ["Disconnect", "Secure against restart", "Verify absence of voltage"],
            "correct_order": [0, 1, 2],
            "explanation": "First disconnect, secure, then verify.",
            "source_ref": "Library/Elektrotechnik/Sicherheit/fuenf-sicherheitsregeln.md",
        },
        {
            "id": "o3",
            "topicId": "it",
            "type": "ordering",
            "prompt": "Order a simple backup workflow.",
            "items": ["Plan", "Run backup", "Test restore"],
            "correct_order": [0, 1, 2],
            "explanation": "Plan first, then run, then test.",
            "source_ref": "Library/IT-Systeme/Dienste/backupverfahren-3-2-1-generationenprinzip.md",
        },
    ]

    for t in topics:
        cur.execute(
            """
            INSERT OR IGNORE INTO topics (id, name, topic_area, parent_id, path, depth)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (t["id"], t["name"], t.get("topic_area", ""), None, t["id"], 0),
        )
    for q in questions:
        cur.execute(
            """
            INSERT INTO questions (id, topic_id, type, data, explanation, source_ref, tags)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                q["id"],
                q["topicId"],
                q["type"],
                json.dumps(q),
                q.get("explanation", ""),
                q.get("source_ref", ""),
                json.dumps(q.get("tags", [])),
            ),
        )
    conn.commit()
    conn.close()


def hash_password(password, salt_b64):
    salt = base64.b64decode(salt_b64)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 200000)
    return base64.b64encode(dk).decode()


def authenticate_user(username, password):
    for user in CONFIG.get("users", []):
        if user["username"] == username:
            hashed = hash_password(password, user["salt"])
            return hashed == user["hash"]
    return False


def create_token(username):
    exp_days = CONFIG.get("token_exp_days", 7)
    payload = {
        "sub": username,
        "exp": datetime.now(timezone.utc) + timedelta(days=exp_days),
    }
    token = jwt.encode(payload, CONFIG["jwt_secret"], algorithm="HS256")
    return token


def get_current_user(request: Request):
    auth = request.headers.get("authorization", "")
    if not auth.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = auth.split(" ", 1)[1]
    try:
        payload = jwt.decode(token, CONFIG["jwt_secret"], algorithms=["HS256"])
        return payload.get("sub")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


@app.on_event("startup")
def on_startup():
    init_db()
    seed_data()


@app.get("/health")
def health():
    print("health check", datetime.now(timezone.utc).isoformat())
    return {"status": "ok"}


@app.post("/auth/login")
def login(data: dict):
    username = data.get("username")
    password = data.get("password")
    if not username or not password:
        raise HTTPException(status_code=400, detail="Missing credentials")
    if not authenticate_user(username, password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token(username)
    return {"token": token, "username": username}


@app.get("/questions")
def get_questions(user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM topics")
    topics = [dict(row) for row in cur.fetchall()]
    cur.execute("SELECT * FROM questions")
    questions = []
    for row in cur.fetchall():
        q = json.loads(row["data"])
        questions.append(q)
    cur.execute("SELECT * FROM packs")
    packs = [dict(row) for row in cur.fetchall()]
    conn.close()
    return {"topics": topics, "questions": questions, "packs": packs}


@app.post("/questions")
def create_question(data: dict, user=Depends(get_current_user)):
    if "id" not in data or "topicId" not in data or "type" not in data:
        raise HTTPException(status_code=400, detail="Missing fields")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM questions WHERE id = ?", (data["id"],))
    if cur.fetchone():
        raise HTTPException(status_code=409, detail="Duplicate id")
    cur.execute(
        "SELECT 1 FROM topics WHERE id = ?",
        (data["topicId"],),
    )
    if not cur.fetchone():
        ensure_topic_entry(
            conn,
            data["topicId"],
            name=data.get("topicName", data["topicId"]),
            topic_area=data.get("topic_area", ""),
            parent_id=data.get("parent_topic_id"),
        )
    cur.execute(
        """
        INSERT INTO questions (id, topic_id, type, data, explanation, source_ref, tags, pack_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            data["id"],
            data["topicId"],
            data["type"],
            json.dumps(data),
            data.get("explanation", ""),
            data.get("source_ref", ""),
            json.dumps(data.get("tags", [])),
            data.get("pack_id", "default_pack"),
        ),
    )
    conn.commit()
    conn.close()
    return {"status": "created", "id": data["id"]}


@app.put("/questions/{question_id}")
def update_question(question_id: str, data: dict, user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data FROM questions WHERE id = ?", (question_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Question not found")
    topic_id = data.get("topicId") or data.get("topic_id")
    if topic_id:
        if not get_topic(conn, topic_id):
            ensure_topic_entry(conn, topic_id, name=data.get("topicName", topic_id))
    stored = json.loads(row["data"])
    stored.update(data)
    stored["id"] = question_id
    if topic_id:
        stored["topicId"] = topic_id
    validate_question_data(stored)
    cur.execute(
        """
        UPDATE questions
        SET topic_id = ?, type = ?, data = ?, explanation = ?, source_ref = ?, tags = ?, pack_id = ?
        WHERE id = ?
        """,
        (
            stored.get("topicId"),
            stored.get("type"),
            json.dumps(stored),
            stored.get("explanation", ""),
            stored.get("source_ref", ""),
            json.dumps(stored.get("tags", [])),
            stored.get("pack_id", "default_pack"),
            question_id,
        ),
    )
    conn.commit()
    conn.close()
    return {"status": "updated", "id": question_id}


@app.delete("/questions/{question_id}")
def delete_question(question_id: str, user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM questions WHERE id = ?", (question_id,))
    conn.commit()
    conn.close()
    return {"status": "deleted", "id": question_id}


def slug_to_title(slug):
    return " ".join([part.capitalize() for part in slug.replace("_", "-").split("-")])


def get_topic(conn, topic_id):
    cur = conn.cursor()
    cur.execute("SELECT * FROM topics WHERE id = ?", (topic_id,))
    row = cur.fetchone()
    return dict(row) if row else None


def compute_topic_path(topic_id, parent):
    if parent and parent.get("path"):
        return f"{parent['path']}/{topic_id}"
    return topic_id


def ensure_topic_entry(conn, topic_id, name=None, topic_area="", parent_id=None):
    cur = conn.cursor()
    cur.execute("SELECT * FROM topics WHERE id = ?", (topic_id,))
    row = cur.fetchone()
    if row:
        return dict(row)
    parent = get_topic(conn, parent_id) if parent_id else None
    path = compute_topic_path(topic_id, parent)
    depth = parent.get("depth", 0) + 1 if parent else 0
    cur.execute(
        "INSERT INTO topics (id, name, topic_area, parent_id, path, depth) VALUES (?, ?, ?, ?, ?, ?)",
        (
            topic_id,
            name or slug_to_title(topic_id),
            topic_area or "",
            parent_id,
            path,
            depth,
        ),
    )
    conn.commit()
    return get_topic(conn, topic_id)


def is_descendant_path(path, potential_parent_path):
    return path == potential_parent_path or path.startswith(f"{potential_parent_path}/")


def move_topic(conn, topic_id, new_parent_id):
    topic = get_topic(conn, topic_id)
    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")
    if new_parent_id == topic_id:
        raise HTTPException(status_code=400, detail="Cannot parent topic to itself")
    parent = get_topic(conn, new_parent_id) if new_parent_id else None
    if parent and is_descendant_path(parent.get("path", ""), topic.get("path", "")):
        raise HTTPException(status_code=400, detail="Cannot move topic into its descendant")
    old_path = topic.get("path") or topic_id
    new_path = compute_topic_path(topic_id, parent)
    depth_delta = (parent.get("depth", 0) + 1 if parent else 0) - (topic.get("depth") or 0)
    cur = conn.cursor()
    cur.execute(
        "UPDATE topics SET parent_id = ?, path = ?, depth = ? WHERE id = ?",
        (new_parent_id, new_path, (topic.get("depth") or 0) + depth_delta, topic_id),
    )
    cur.execute("SELECT id, path, depth FROM topics WHERE path LIKE ?", (f"{old_path}/%",))
    descendants = [dict(row) for row in cur.fetchall()]
    for desc in descendants:
        new_desc_path = desc["path"].replace(f"{old_path}/", f"{new_path}/", 1)
        new_desc_depth = (desc.get("depth") or 0) + depth_delta
        cur.execute(
            "UPDATE topics SET path = ?, depth = ? WHERE id = ?",
            (new_desc_path, new_desc_depth, desc["id"]),
        )
    conn.commit()


def normalize_question(input_q, topic_map, errors):
    if not input_q or "id" not in input_q or "type" not in input_q:
        errors.append("Question missing id or type.")
        return None
    type_map = {"guessword": "guess", "explainterm": "explain"}
    q_type = type_map.get(input_q["type"], input_q["type"])
    topic_slug = input_q.get("topic_slug") or input_q.get("topicId")
    if not topic_slug:
        errors.append(f"Question {input_q['id']} missing topic_slug.")
        return None
    if topic_slug not in topic_map:
        topic_map[topic_slug] = {
            "id": topic_slug,
            "name": slug_to_title(topic_slug),
            "topic_area": "",
        }
    stem_id = input_q.get("stem_id") or input_q.get("stemId")
    stem_text = (
        input_q.get("stem_text")
        or input_q.get("stemText")
        or input_q.get("stem")
        or input_q.get("prompt")
        or input_q.get("text")
    )
    prompt = input_q.get("prompt") or input_q.get("statement") or input_q.get("text")
    if not prompt:
        errors.append(f"Question {input_q['id']} missing prompt.")
        return None

    base = {
        "id": input_q["id"],
        "topicId": topic_slug,
        "type": q_type,
        "prompt": prompt,
        "explanation": input_q.get("explanation", ""),
        "source_ref": input_q.get("source_ref", "internal:import"),
        "tags": input_q.get("tags", []),
        "stem_id": stem_id,
        "stem_text": stem_text,
    }

    if q_type == "single":
        if not isinstance(input_q.get("options"), list) or not isinstance(input_q.get("correct"), list):
            errors.append(f"Single {input_q['id']} missing options/correct.")
            return None
        return {**base, "options": input_q["options"], "correctIndex": input_q["correct"][0]}
    if q_type == "multi":
        if not isinstance(input_q.get("options"), list) or not isinstance(input_q.get("correct"), list):
            errors.append(f"Multi {input_q['id']} missing options/correct.")
            return None
        return {**base, "options": input_q["options"], "correctIndexes": input_q["correct"]}
    if q_type == "truefalse":
        if not isinstance(input_q.get("correct"), bool):
            errors.append(f"TrueFalse {input_q['id']} missing boolean correct.")
            return None
        return {**base, "correctBoolean": input_q["correct"]}
    if q_type == "fillblank":
        if not isinstance(input_q.get("answers"), list):
            errors.append(f"Fillblank {input_q['id']} missing answers.")
            return None
        return {**base, "answers": input_q["answers"]}
    if q_type == "matching":
        left = input_q.get("left")
        right = input_q.get("right")
        pairs = input_q.get("pairs")
        if not isinstance(left, list) or not isinstance(right, list):
            errors.append(f"Matching {input_q['id']} missing left/right.")
            return None
        if isinstance(pairs, list) and len(pairs) > 0:
            mapped = [{"left": left[p[0]], "right": right[p[1]]} for p in pairs]
        elif len(left) == len(right):
            mapped = [{"left": l, "right": right[i]} for i, l in enumerate(left)]
        else:
            errors.append(f"Matching {input_q['id']} missing pairs.")
            return None
        return {**base, "pairs": mapped}
    if q_type == "ordering":
        if not isinstance(input_q.get("items"), list) or not isinstance(input_q.get("correct_order"), list):
            errors.append(f"Ordering {input_q['id']} missing items/correct_order.")
            return None
        return {**base, "items": input_q["items"], "correct_order": input_q["correct_order"]}
    if q_type == "guess":
        if not isinstance(input_q.get("answers"), list):
            errors.append(f"Guess {input_q['id']} missing answers.")
            return None
        return {**base, "expectedAnswers": input_q["answers"]}
    if q_type == "explain":
        return {**base, "expectedAnswer": ", ".join(input_q.get("keywords", []))}
    if q_type == "exam":
        return {**base, "expectedAnswer": input_q.get("answer_key", "")}
    errors.append(f"Unknown type {q_type} for {input_q['id']}.")
    return None


def validate_question_data(question):
    q_type = question.get("type")
    if not q_type:
        raise HTTPException(status_code=400, detail="Missing type")
    if q_type == "single":
        options = question.get("options")
        correct = question.get("correctIndex")
        if not isinstance(options, list) or len(options) < 2:
            raise HTTPException(status_code=400, detail="Single requires options")
        if correct is None or not isinstance(correct, int) or correct < 0 or correct >= len(options):
            raise HTTPException(status_code=400, detail="Single requires valid correctIndex")
        return
    if q_type == "multi":
        options = question.get("options")
        correct = question.get("correctIndexes")
        if not isinstance(options, list) or len(options) < 2:
            raise HTTPException(status_code=400, detail="Multi requires options")
        if not isinstance(correct, list) or len(correct) == 0:
            raise HTTPException(status_code=400, detail="Multi requires correctIndexes")
        if any((not isinstance(idx, int) or idx < 0 or idx >= len(options)) for idx in correct):
            raise HTTPException(status_code=400, detail="Multi requires valid correctIndexes")
        return
    if q_type == "truefalse":
        if not isinstance(question.get("correctBoolean"), bool):
            raise HTTPException(status_code=400, detail="Truefalse requires correctBoolean")
        return
    if q_type == "fillblank":
        answers = question.get("answers") or question.get("expectedAnswers")
        if not isinstance(answers, list) or len(answers) == 0:
            raise HTTPException(status_code=400, detail="Fillblank requires answers")
        return
    if q_type == "matching":
        pairs = question.get("pairs")
        if not isinstance(pairs, list) or len(pairs) == 0:
            raise HTTPException(status_code=400, detail="Matching requires pairs")
        return
    if q_type == "ordering":
        items = question.get("items")
        correct = question.get("correct_order")
        if not isinstance(items, list) or len(items) == 0:
            raise HTTPException(status_code=400, detail="Ordering requires items")
        if not isinstance(correct, list) or len(correct) != len(items):
            raise HTTPException(status_code=400, detail="Ordering requires correct_order")
        return
    if q_type == "guess":
        answers = question.get("expectedAnswers")
        if not isinstance(answers, list) or len(answers) == 0:
            raise HTTPException(status_code=400, detail="Guess requires expectedAnswers")
        return
    if q_type == "explain":
        expected = question.get("expectedAnswer")
        if not isinstance(expected, str):
            raise HTTPException(status_code=400, detail="Explain requires expectedAnswer")
        return
    if q_type == "exam":
        expected = question.get("expectedAnswer")
        if not isinstance(expected, str):
            raise HTTPException(status_code=400, detail="Exam requires expectedAnswer")
        return
    if q_type == "calc_value":
        ok, msg = validate_calc_value_payload(question.get("payload", {}))
        if not ok:
            raise HTTPException(status_code=400, detail=msg)
        return
    if q_type == "calc_multi":
        ok, msg = validate_calc_multi_payload(question.get("payload", {}))
        if not ok:
            raise HTTPException(status_code=400, detail=msg)
        return
    if q_type == "hotspot_svg":
        ok, msg = validate_hotspot_payload(question.get("payload", {}))
        if not ok:
            raise HTTPException(status_code=400, detail=msg)
        return
    if q_type == "troubleshoot_flow":
        ok, msg = validate_troubleshoot_payload(question.get("payload", {}))
        if not ok:
            raise HTTPException(status_code=400, detail=msg)
        return


def get_question_record(conn, question_id):
    cur = conn.cursor()
    cur.execute("SELECT * FROM questions WHERE id = ?", (question_id,))
    row = cur.fetchone()
    if not row:
        return None
    data = json.loads(row["data"])
    return data


def compare_set(a, b):
    return set(a or []) == set(b or [])


def grade_question(question, answer):
    q_type = question.get("type")
    payload = question.get("payload", {})
    if q_type == "single":
        selected = answer.get("selected")
        correct = selected == question.get("correctIndex")
        expected = question["options"][question["correctIndex"]]
        return True, correct, expected, None
    if q_type == "multi":
        selected = answer.get("selected") or []
        correct = compare_set(selected, question.get("correctIndexes", []))
        expected = ", ".join(
            [question["options"][i] for i in question.get("correctIndexes", [])]
        )
        return True, correct, expected, None
    if q_type == "truefalse":
        selected = answer.get("selected")
        correct = bool(selected) == bool(question.get("correctBoolean"))
        expected = "True" if question.get("correctBoolean") else "False"
        return True, correct, expected, None
    if q_type == "matching":
        selected = answer.get("selected") or []
        correct = True
        pairs = question.get("pairs") or []
        if len(selected) != len(pairs):
            correct = False
        else:
            for idx, pair in enumerate(pairs):
                if selected[idx] != pair.get("right"):
                    correct = False
                    break
        expected = "; ".join([f"{p['left']} -> {p['right']}" for p in pairs])
        return True, correct, expected, None
    if q_type == "ordering":
        selected = answer.get("selected") or []
        correct = selected == question.get("correct_order")
        expected = ", ".join([str(i + 1) for i in question.get("correct_order", [])])
        return True, correct, expected, None
    if q_type == "fillblank":
        selected = answer.get("selected") or []
        blanks = question.get("answers") or question.get("expectedAnswers") or []
        if len(selected) != len(blanks):
            return True, False, "Fill all blanks", None
        case_sensitive = payload.get("case_sensitive")
        if case_sensitive is None:
            case_sensitive = False
        correct = True
        for idx, val in enumerate(selected):
            accepted = blanks[idx]
            if not isinstance(accepted, list):
                accepted = [accepted]
            if case_sensitive:
                ok = val in accepted
            else:
                ok = str(val).lower() in [str(a).lower() for a in accepted]
            if not ok:
                correct = False
                break
        expected = " | ".join(
            ["/".join([str(a) for a in (b if isinstance(b, list) else [b])]) for b in blanks]
        )
        return True, correct, expected, None
    if q_type == "guess":
        value = answer.get("text", "")
        accepted = question.get("expectedAnswers") or []
        correct = str(value).lower() in [str(a).lower() for a in accepted]
        expected = ", ".join([str(a) for a in accepted])
        return True, correct, expected, None
    if q_type in ("explain", "exam"):
        expected = question.get("expectedAnswer") or ""
        return False, None, expected, None
    if q_type == "calc_value":
        expected_value = payload.get("expected_value")
        expected_unit = payload.get("expected_unit")
        accept_units = payload.get("accept_units") or [expected_unit]
        rounding = payload.get("rounding_decimals", 2)
        tol = payload.get("tolerance") or {"mode": "absolute", "value": 0}
        if "raw" in answer:
            user_value, user_unit, raw_unit, err = parse_value_unit(answer.get("raw"), None)
        else:
            user_value, user_unit, raw_unit, err = parse_value_unit(
                answer.get("value"), answer.get("unit")
            )
        if err:
            return True, False, "Invalid value/unit", err
        expected_base, expected_unit_base, _, err2 = parse_value_unit(
            expected_value, expected_unit
        )
        if err2:
            return True, False, "Invalid expected unit", err2
        if raw_unit and accept_units:
            allowed_units = [str(u).lower() for u in accept_units if u]
            if str(raw_unit).lower() not in allowed_units:
                return True, False, "Unit not accepted", "Unit not accepted"
            return True, False, "Unit not accepted", "Unit not accepted"
        if expected_unit_base and user_unit and expected_unit_base != user_unit:
            return True, False, "Wrong unit", "Wrong unit"
        user_value = round_value(user_value, rounding)
        expected_base = round_value(expected_base, rounding)
        diff = abs(user_value - expected_base)
        mode = tol.get("mode", "absolute")
        tol_val = float(tol.get("value", 0))
        if mode == "relative":
            allowed = abs(expected_base) * tol_val
        else:
            allowed = tol_val
        correct = diff <= allowed
        expected = f"{expected_base} {expected_unit}"
        return True, correct, expected, None
    if q_type == "calc_multi":
        fields = payload.get("fields") or []
        answers = payload.get("answers") or {}
        input_fields = answer.get("fields") or {}
        all_correct = True
        for field in fields:
            fid = field.get("id")
            expected = answers.get(fid)
            if fid not in input_fields or expected is None:
                all_correct = False
                continue
            user = input_fields.get(fid, {})
            user_value, user_unit, raw_unit, err = parse_value_unit(
                user.get("value"), user.get("unit")
            )
            if err:
                all_correct = False
                continue
            expected_value, expected_unit, _, err2 = parse_value_unit(
                expected.get("value"), expected.get("unit")
            )
            if err2:
                all_correct = False
                continue
            if expected_unit and user_unit and expected_unit != user_unit:
                all_correct = False
                continue
            decimals = field.get("decimals", 2)
            tol = field.get("tolerance") or {"mode": "absolute", "value": 0}
            user_value = round_value(user_value, decimals)
            expected_value = round_value(expected_value, decimals)
            diff = abs(user_value - expected_value)
            mode = tol.get("mode", "absolute")
            tol_val = float(tol.get("value", 0))
            if mode == "relative":
                allowed = abs(expected_value) * tol_val
            else:
                allowed = tol_val
            if diff > allowed:
                all_correct = False
        expected_text = ", ".join([f"{k}: {v.get('value')} {v.get('unit','')}".strip() for k, v in answers.items()])
        return True, all_correct, expected_text, None
    if q_type == "hotspot_svg":
        correct = compare_set(answer.get("selected") or [], payload.get("correct") or [])
        expected = ", ".join([str(x) for x in payload.get("correct") or []])
        return True, correct, expected, None
    if q_type == "troubleshoot_flow":
        final_node = answer.get("final_node")
        correct = final_node == payload.get("success_node")
        expected = payload.get("success_node")
        return True, correct, expected, None
    return True, False, "Unknown type", "Unknown type"


def import_quiztab_v2(pack, replace_duplicates):
    errors = []
    topics_input = pack.get("topics", [])
    questions_input = pack.get("questions", [])
    if not isinstance(topics_input, list) or not isinstance(questions_input, list):
        raise HTTPException(status_code=400, detail="Invalid quiztab pack")

    topic_map = build_topic_map(topics_input, errors)
    pack_id = get_pack_id(pack)

    normalized = []
    summary = {"total": 0, "valid": 0, "invalid": 0, "types": {}, "topics": {}}
    for q in questions_input:
        summary["total"] += 1
        nq = normalize_quiztab_question(q, topic_map, errors, pack_id)
        if nq:
            validate_question_data(nq)
            normalized.append(nq)
            summary["valid"] += 1
            summary["types"][nq["type"]] = summary["types"].get(nq["type"], 0) + 1
            summary["topics"][nq["topicId"]] = summary["topics"].get(nq["topicId"], 0) + 1
        else:
            summary["invalid"] += 1

    conn = get_db()
    cur = conn.cursor()
    added = 0
    skipped = 0
    replaced = 0

    for t in topic_map.values():
        cur.execute("SELECT 1 FROM topics WHERE id = ?", (t["id"],))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO topics (id, name, topic_area, parent_id, path, depth) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    t["id"],
                    t["name"],
                    t.get("topic_area", ""),
                    t.get("parent_id"),
                    t.get("path") or t["id"],
                    t.get("depth", 0),
                ),
            )

    upsert_pack(conn, pack_id, pack)

    for q in normalized:
        cur.execute("SELECT 1 FROM questions WHERE id = ?", (q["id"],))
        exists = cur.fetchone() is not None
        if exists and not replace_duplicates:
            skipped += 1
            continue
        if exists and replace_duplicates:
            cur.execute(
                """
                UPDATE questions
                SET topic_id = ?, type = ?, data = ?, explanation = ?, source_ref = ?, tags = ?, pack_id = ?
                WHERE id = ?
                """,
                (
                    q["topicId"],
                    q["type"],
                    json.dumps(q),
                    q.get("explanation", ""),
                    q.get("source_ref", ""),
                    json.dumps(q.get("tags", [])),
                    q.get("pack_id"),
                    q["id"],
                ),
            )
            replaced += 1
            continue
        cur.execute(
            """
            INSERT INTO questions (id, topic_id, type, data, explanation, source_ref, tags, pack_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                q["id"],
                q["topicId"],
                q["type"],
                json.dumps(q),
                q.get("explanation", ""),
                q.get("source_ref", ""),
                json.dumps(q.get("tags", [])),
                q.get("pack_id"),
            ),
        )
        added += 1

    conn.commit()
    conn.close()

    return {
        "summary": summary,
        "errors": errors[:5],
        "added": added,
        "skipped": skipped,
        "replaced": replaced,
        "pack_id": pack_id,
    }


@app.post("/questions/import")
def import_questions(data: dict, replace_duplicates: bool = False, user=Depends(get_current_user)):
    schema = data.get("schema")
    if schema == "quiztab-questionpack-v2":
        return import_quiztab_v2(data, replace_duplicates)
    if schema not in ("ap2-questionpack-v1", "ap2-questionpack-v2"):
        raise HTTPException(status_code=400, detail="Invalid schema")
    topics_input = data.get("topics", [])
    questions_input = data.get("questions", [])
    stems_input = data.get("stems", [])
    errors = []
    topic_map = build_topic_map(topics_input, errors)

    # Expand stems into questions.
    for stem in stems_input:
        stem_id = stem.get("stem_id") or stem.get("id")
        stem_text = stem.get("stem_text") or stem.get("text") or stem.get("prompt")
        stem_topic = stem.get("topic_id") or stem.get("topic_slug") or stem.get("topicId")
        if not stem_id or not stem_text:
            errors.append("Stem missing stem_id or stem_text.")
            continue
        variants = stem.get("variants", [])
        if not isinstance(variants, list) or len(variants) == 0:
            errors.append(f"Stem {stem_id} missing variants.")
            continue
        for variant in variants:
            if not isinstance(variant, dict):
                errors.append(f"Stem {stem_id} variant is not an object.")
                continue
            variant_id = variant.get("id")
            if not variant_id:
                errors.append(f"Stem {stem_id} variant missing id.")
                continue
            variant_topic = (
                variant.get("topic_slug")
                or variant.get("topic_id")
                or variant.get("topicId")
                or stem_topic
            )
            questions_input.append(
                {
                    **variant,
                    "id": variant_id,
                    "stem_id": stem_id,
                    "stem_text": stem_text,
                    "topic_slug": variant_topic,
                    "prompt": variant.get("prompt") or stem_text,
                }
            )

    normalized = []
    summary = {"total": 0, "valid": 0, "invalid": 0, "types": {}, "topics": {}}
    for q in questions_input:
        summary["total"] += 1
        nq = normalize_question(q, topic_map, errors)
        if nq:
            normalized.append(nq)
            summary["valid"] += 1
            summary["types"][nq["type"]] = summary["types"].get(nq["type"], 0) + 1
            summary["topics"][nq["topicId"]] = summary["topics"].get(nq["topicId"], 0) + 1
        else:
            summary["invalid"] += 1

    conn = get_db()
    cur = conn.cursor()
    added = 0
    skipped = 0
    replaced = 0

    for t in topic_map.values():
        cur.execute("SELECT 1 FROM topics WHERE id = ?", (t["id"],))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO topics (id, name, topic_area, parent_id, path, depth) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    t["id"],
                    t["name"],
                    t.get("topic_area", ""),
                    t.get("parent_id"),
                    t.get("path") or t["id"],
                    t.get("depth", 0),
                ),
            )

    for q in normalized:
        cur.execute("SELECT 1 FROM questions WHERE id = ?", (q["id"],))
        exists = cur.fetchone() is not None
        if exists and not replace_duplicates:
            skipped += 1
            continue
        if exists and replace_duplicates:
            cur.execute(
                "UPDATE questions SET topic_id = ?, type = ?, data = ?, explanation = ?, source_ref = ?, tags = ? WHERE id = ?",
                (
                    q["topicId"],
                    q["type"],
                    json.dumps(q),
                    q.get("explanation", ""),
                    q.get("source_ref", ""),
                    json.dumps(q.get("tags", [])),
                    q["id"],
                ),
            )
            replaced += 1
            continue
        cur.execute(
            "INSERT INTO questions (id, topic_id, type, data, explanation, source_ref, tags) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                q["id"],
                q["topicId"],
                q["type"],
                json.dumps(q),
                q.get("explanation", ""),
                q.get("source_ref", ""),
                json.dumps(q.get("tags", [])),
            ),
        )
        added += 1

    conn.commit()
    conn.close()

    return {
        "summary": summary,
        "errors": errors[:5],
        "added": added,
        "skipped": skipped,
        "replaced": replaced,
    }


@app.get("/questions/export")
def export_questions(user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM topics")
    topics = [dict(row) for row in cur.fetchall()]
    cur.execute("SELECT * FROM questions")
    questions = [json.loads(row["data"]) for row in cur.fetchall()]
    conn.close()
    pack = {
        "schema": "ap2-questionpack-v2",
        "meta": {
            "title": "AP2 Export",
            "generated_by": "AP2 Trainer",
            "created_at": datetime.now().strftime("%Y-%m-%d"),
        },
        "topics": [
            {
                "slug": t["id"],
                "title": t["name"],
                "topic_area": t.get("topic_area", ""),
                "parent_topic_id": t.get("parent_id"),
                "path": t.get("path") or t["id"],
                "depth": t.get("depth") if t.get("depth") is not None else 0,
            }
            for t in topics
        ],
        "stems": [],
        "questions": [],
    }
    stems = {}
    for q in questions:
        stem_id = q.get("stem_id")
        if stem_id:
            if stem_id not in stems:
                stems[stem_id] = {
                    "stem_id": stem_id,
                    "stem_text": q.get("stem_text") or q.get("prompt") or "",
                    "topic_id": q.get("topicId"),
                    "variants": [],
                }
            stems[stem_id]["variants"].append(q)
            continue
        base = {
            "id": q["id"],
            "topic_slug": q["topicId"],
            "type": q["type"],
            "prompt": q.get("prompt"),
            "explanation": q.get("explanation", ""),
            "source_ref": q.get("source_ref", ""),
            "tags": q.get("tags", []),
        }
        if q["type"] == "single":
            pack["questions"].append({**base, "options": q["options"], "correct": [q["correctIndex"]]})
        elif q["type"] == "multi":
            pack["questions"].append({**base, "options": q["options"], "correct": q["correctIndexes"]})
        elif q["type"] == "truefalse":
            pack["questions"].append({**base, "correct": q["correctBoolean"]})
        elif q["type"] == "fillblank":
            pack["questions"].append({**base, "text": q.get("prompt"), "answers": q.get("answers", q.get("expectedAnswers", []))})
        elif q["type"] == "matching":
            left = [p["left"] for p in q["pairs"]]
            right = [p["right"] for p in q["pairs"]]
            pairs = [[i, i] for i in range(len(q["pairs"]))]
            pack["questions"].append({**base, "left": left, "right": right, "pairs": pairs})
        elif q["type"] == "ordering":
            pack["questions"].append({**base, "items": q["items"], "correct_order": q["correct_order"]})
        elif q["type"] == "guess":
            pack["questions"].append({**base, "type": "guessword", "answers": q.get("expectedAnswers", [])})
        elif q["type"] == "explain":
            pack["questions"].append({**base, "type": "explainterm", "keywords": q.get("expectedAnswer", "").split(",")})
        elif q["type"] == "exam":
            pack["questions"].append({**base, "answer_key": q.get("expectedAnswer", ""), "grading": "self"})
        else:
            pack["questions"].append(base)
    for stem in stems.values():
        variants = []
        for q in stem["variants"]:
            base = {
                "id": q["id"],
                "type": q["type"],
                "prompt": q.get("prompt"),
                "explanation": q.get("explanation", ""),
                "source_ref": q.get("source_ref", ""),
                "tags": q.get("tags", []),
            }
            if q["type"] == "single":
                variants.append({**base, "options": q["options"], "correct": [q["correctIndex"]]})
            elif q["type"] == "multi":
                variants.append({**base, "options": q["options"], "correct": q["correctIndexes"]})
            elif q["type"] == "truefalse":
                variants.append({**base, "correct": q["correctBoolean"]})
            elif q["type"] == "fillblank":
                variants.append({**base, "text": q.get("prompt"), "answers": q.get("answers", q.get("expectedAnswers", []))})
            elif q["type"] == "matching":
                left = [p["left"] for p in q["pairs"]]
                right = [p["right"] for p in q["pairs"]]
                pairs = [[i, i] for i in range(len(q["pairs"]))]
                variants.append({**base, "left": left, "right": right, "pairs": pairs})
            elif q["type"] == "ordering":
                variants.append({**base, "items": q["items"], "correct_order": q["correct_order"]})
            elif q["type"] == "guess":
                variants.append({**base, "type": "guessword", "answers": q.get("expectedAnswers", [])})
            elif q["type"] == "explain":
                variants.append({**base, "type": "explainterm", "keywords": q.get("expectedAnswer", "").split(",")})
            elif q["type"] == "exam":
                variants.append({**base, "answer_key": q.get("expectedAnswer", ""), "grading": "self"})
            else:
                variants.append(base)
        pack["stems"].append({**stem, "variants": variants})
    return pack


@app.get("/packs/export")
def export_pack(schema: str = "quiztab-questionpack-v2", pack_id: str = "default_pack", user=Depends(get_current_user)):
    if schema != "quiztab-questionpack-v2":
        raise HTTPException(status_code=400, detail="Unsupported schema")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM packs WHERE id = ?", (pack_id,))
    pack_row = cur.fetchone()
    if pack_row:
        meta = json.loads(pack_row["meta"]) if pack_row["meta"] else {}
        settings = json.loads(pack_row["settings"]) if pack_row["settings"] else {}
        methods = json.loads(pack_row["methods"]) if pack_row["methods"] else []
        assets = json.loads(pack_row["assets"]) if pack_row["assets"] else []
    else:
        meta = {"title": "QuizTab Export", "created_at": datetime.now().strftime("%Y-%m-%d")}
        settings = {}
        methods = []
        assets = []

    cur.execute("SELECT * FROM topics")
    topics = [dict(row) for row in cur.fetchall()]
    cur.execute("SELECT data FROM questions WHERE pack_id = ?", (pack_id,))
    questions = [json.loads(row["data"]) for row in cur.fetchall()]
    conn.close()

    pack = {
        "schema": "quiztab-questionpack-v2",
        "meta": meta,
        "settings": settings,
        "topics": [
            {
                "slug": t["id"],
                "title": t["name"],
                "topic_area": t.get("topic_area", ""),
                "parent_topic_id": t.get("parent_id"),
                "path": t.get("path") or t["id"],
                "depth": t.get("depth") if t.get("depth") is not None else 0,
            }
            for t in topics
        ],
        "methods": methods,
        "assets": assets,
        "questions": [
            {
                "id": q.get("id"),
                "topic_slug": q.get("topicId"),
                "method_id": q.get("method_id") or q.get("type"),
                "difficulty": q.get("difficulty", ""),
                "prompt": q.get("prompt"),
                "payload": q.get("payload", {}),
                "support": q.get("support", {}),
                "solution": q.get("solution", {}),
                "source_ref": q.get("source_ref", ""),
                "tags": q.get("tags", []),
                "variants": q.get("variants", []),
                "randomization": q.get("randomization"),
            }
            for q in questions
        ],
    }
    return pack


def parse_frontmatter(text):
    lines = text.splitlines()
    header = {}
    body_start = 0
    if lines and lines[0].strip().startswith("{"):
        buf = []
        for idx, line in enumerate(lines):
            buf.append(line)
            if line.strip().endswith("}"):
                body_start = idx + 1
                try:
                    header = json.loads("\n".join(buf))
                except json.JSONDecodeError:
                    header = {}
                break
    else:
        for idx, line in enumerate(lines):
            if not line.strip():
                body_start = idx + 1
                break
            if ":" in line:
                key, val = line.split(":", 1)
                header[key.strip()] = val.strip()
        else:
            body_start = len(lines)
    return header, "\n".join(lines[body_start:])


def parse_sectioned_body(body):
    sections = {"prompt": "", "payload": "", "support": "", "solution": ""}
    current = None
    for line in body.splitlines():
        key = line.strip().lower()
        if key.startswith("prompt:"):
            current = "prompt"
            sections[current] = line.split(":", 1)[1].strip()
            continue
        if key.startswith("payload:"):
            current = "payload"
            sections[current] = line.split(":", 1)[1].strip()
            continue
        if key.startswith("support:"):
            current = "support"
            sections[current] = line.split(":", 1)[1].strip()
            continue
        if key.startswith("solution:"):
            current = "solution"
            sections[current] = line.split(":", 1)[1].strip()
            continue
        if current:
            sections[current] += ("\n" if sections[current] else "") + line
    return sections


@app.post("/packs/import_markdown_zip")
def import_markdown_zip(file: UploadFile = File(...), user=Depends(get_current_user)):
    raw = file.file.read()
    zf = zipfile.ZipFile(io.BytesIO(raw))
    questions = []
    topics = {}
    for name in zf.namelist():
        if not name.lower().endswith(".md"):
            continue
        content = zf.read(name).decode("utf-8", errors="ignore")
        header, body = parse_frontmatter(content)
        sections = parse_sectioned_body(body)
        qid = header.get("id") or f"q_{uuid.uuid4().hex}"
        topic_slug = header.get("topic_slug") or "general"
        method_id = header.get("method_id") or "multi"
        difficulty = header.get("difficulty") or ""
        tags = [t.strip() for t in (header.get("tags") or "").split(",") if t.strip()]
        source_ref = header.get("source_ref") or "internal:import"

        def parse_json_block(text):
            if not text.strip():
                return {}
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                return {}

        payload = parse_json_block(sections.get("payload", ""))
        support = parse_json_block(sections.get("support", ""))
        solution = parse_json_block(sections.get("solution", ""))
        prompt = sections.get("prompt", "").strip()
        questions.append(
            {
                "id": qid,
                "topic_slug": topic_slug,
                "method_id": method_id,
                "difficulty": difficulty,
                "prompt": prompt,
                "payload": payload,
                "support": support,
                "solution": solution,
                "source_ref": source_ref,
                "tags": tags,
            }
        )
        topics[topic_slug] = {"slug": topic_slug, "title": slug_to_title(topic_slug)}

    pack = {
        "schema": "quiztab-questionpack-v2",
        "meta": {
            "title": "Markdown Import",
            "created_at": datetime.now().strftime("%Y-%m-%d"),
            "source": "markdown_zip",
        },
        "settings": {},
        "topics": list(topics.values()),
        "methods": [],
        "assets": [],
        "questions": questions,
    }
    return import_quiztab_v2(pack, replace_duplicates=False)


@app.post("/packs/import_excel")
def import_excel(file: UploadFile = File(...), user=Depends(get_current_user)):
    import openpyxl

    raw = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    if "questions" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Missing sheet 'questions'")
    ws = wb["questions"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers) if name}
    questions = []
    topics = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid = row[idx.get("id")] if "id" in idx else None
        topic_slug = row[idx.get("topic_slug")] if "topic_slug" in idx else "general"
        method_id = row[idx.get("method_id")] if "method_id" in idx else "multi"
        prompt = row[idx.get("prompt")] if "prompt" in idx else ""
        payload_raw = row[idx.get("payload_json")] if "payload_json" in idx else "{}"
        solution_raw = row[idx.get("solution_json")] if "solution_json" in idx else "{}"
        tags_raw = row[idx.get("tags_csv")] if "tags_csv" in idx else ""
        source_ref = row[idx.get("source_ref")] if "source_ref" in idx else "internal:import"
        try:
            payload = json.loads(payload_raw or "{}")
        except json.JSONDecodeError:
            payload = {}
        try:
            solution = json.loads(solution_raw or "{}")
        except json.JSONDecodeError:
            solution = {}
        tags = [t.strip() for t in str(tags_raw or "").split(",") if t.strip()]
        questions.append(
            {
                "id": qid or f"q_{uuid.uuid4().hex}",
                "topic_slug": topic_slug or "general",
                "method_id": method_id or "multi",
                "difficulty": "",
                "prompt": prompt or "",
                "payload": payload,
                "support": {},
                "solution": solution,
                "source_ref": source_ref,
                "tags": tags,
            }
        )
        topics[topic_slug] = {"slug": topic_slug, "title": slug_to_title(topic_slug)}

    pack = {
        "schema": "quiztab-questionpack-v2",
        "meta": {
            "title": "Excel Import",
            "created_at": datetime.now().strftime("%Y-%m-%d"),
            "source": "excel",
        },
        "settings": {},
        "topics": list(topics.values()),
        "methods": [],
        "assets": [],
        "questions": questions,
    }
    return import_quiztab_v2(pack, replace_duplicates=False)


@app.get("/topics")
def get_topics(user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM topics")
    topics = [dict(row) for row in cur.fetchall()]
    conn.close()
    return {"topics": topics}


@app.post("/topics")
def create_topic(data: dict, user=Depends(get_current_user)):
    topic_id = data.get("id") or data.get("slug")
    name = data.get("name") or data.get("title")
    parent_id = data.get("parent_id") or data.get("parent_topic_id")
    topic_area = data.get("topic_area", "")
    if not topic_id:
        raise HTTPException(status_code=400, detail="Missing topic id")
    conn = get_db()
    if get_topic(conn, topic_id):
        conn.close()
        raise HTTPException(status_code=409, detail="Topic already exists")
    parent = get_topic(conn, parent_id) if parent_id else None
    if parent_id and not parent:
        conn.close()
        raise HTTPException(status_code=400, detail="Parent topic not found")
    path = compute_topic_path(topic_id, parent)
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM topics WHERE path = ?", (path,))
    if cur.fetchone():
        conn.close()
        raise HTTPException(status_code=409, detail="Topic path already exists")
    depth = parent.get("depth", 0) + 1 if parent else 0
    cur.execute(
        "INSERT INTO topics (id, name, topic_area, parent_id, path, depth) VALUES (?, ?, ?, ?, ?, ?)",
        (topic_id, name or slug_to_title(topic_id), topic_area, parent_id, path, depth),
    )
    conn.commit()
    conn.close()
    return {"status": "created", "id": topic_id}


@app.put("/topics/{topic_id}")
def update_topic(topic_id: str, data: dict, user=Depends(get_current_user)):
    conn = get_db()
    topic = get_topic(conn, topic_id)
    if not topic:
        conn.close()
        raise HTTPException(status_code=404, detail="Topic not found")
    name = data.get("name") or data.get("title")
    new_parent_id = data.get("parent_id") or data.get("parent_topic_id")
    if new_parent_id is not None and new_parent_id != topic.get("parent_id"):
        move_topic(conn, topic_id, new_parent_id)
    if name:
        cur = conn.cursor()
        cur.execute("UPDATE topics SET name = ? WHERE id = ?", (name, topic_id))
        conn.commit()
    conn.close()
    return {"status": "updated", "id": topic_id}


@app.delete("/topics/{topic_id}")
def delete_topic(topic_id: str, mode: str = "delete", user=Depends(get_current_user)):
    conn = get_db()
    topic = get_topic(conn, topic_id)
    if not topic:
        conn.close()
        raise HTTPException(status_code=404, detail="Topic not found")
    cur = conn.cursor()
    if mode == "reassign":
        parent_id = topic.get("parent_id")
        if not parent_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Cannot reassign root topic")
        cur.execute("SELECT id FROM topics WHERE parent_id = ?", (topic_id,))
        child_ids = [row["id"] for row in cur.fetchall()]
        for child_id in child_ids:
            move_topic(conn, child_id, parent_id)
        cur.execute("UPDATE questions SET topic_id = ? WHERE topic_id = ?", (parent_id, topic_id))
        cur.execute("DELETE FROM topics WHERE id = ?", (topic_id,))
        conn.commit()
        conn.close()
        return {"status": "deleted", "mode": "reassign", "id": topic_id}
    if mode != "delete":
        conn.close()
        raise HTTPException(status_code=400, detail="Invalid delete mode")
    path = topic.get("path") or topic_id
    cur.execute("SELECT id FROM topics WHERE id = ? OR path LIKE ?", (topic_id, f"{path}/%"))
    delete_ids = [row["id"] for row in cur.fetchall()]
    if delete_ids:
        placeholders = ",".join(["?"] * len(delete_ids))
        cur.execute(f"DELETE FROM questions WHERE topic_id IN ({placeholders})", delete_ids)
        cur.execute(f"DELETE FROM topics WHERE id IN ({placeholders})", delete_ids)
    conn.commit()
    conn.close()
    return {"status": "deleted", "mode": "delete", "id": topic_id, "topics_deleted": len(delete_ids)}


def convert_question_type(question, target_type):
    current = question.get("type")
    if current == target_type:
        return question
    if current == "single" and target_type == "multi":
        question["type"] = "multi"
        question["correctIndexes"] = [question.get("correctIndex", 0)]
        question.pop("correctIndex", None)
        return question
    if current == "multi" and target_type == "single":
        correct = question.get("correctIndexes", [])
        question["type"] = "single"
        question["correctIndex"] = correct[0] if correct else 0
        question.pop("correctIndexes", None)
        return question
    raise HTTPException(status_code=400, detail="Unsupported type conversion")


@app.post("/questions/bulk")
def bulk_questions(data: dict, user=Depends(get_current_user)):
    action = data.get("action")
    ids = data.get("ids") or []
    if not isinstance(ids, list) or len(ids) == 0:
        raise HTTPException(status_code=400, detail="Missing ids")
    conn = get_db()
    cur = conn.cursor()
    if action == "delete":
        placeholders = ",".join(["?"] * len(ids))
        cur.execute(f"DELETE FROM questions WHERE id IN ({placeholders})", ids)
        conn.commit()
        conn.close()
        return {"status": "deleted", "count": cur.rowcount}
    if action == "move":
        topic_id = data.get("topic_id")
        if not topic_id:
            conn.close()
            raise HTTPException(status_code=400, detail="Missing topic_id")
        if not get_topic(conn, topic_id):
            conn.close()
            raise HTTPException(status_code=400, detail="Topic not found")
        placeholders = ",".join(["?"] * len(ids))
        cur.execute(f"SELECT id, data FROM questions WHERE id IN ({placeholders})", ids)
        rows = [dict(row) for row in cur.fetchall()]
        for row in rows:
            q = json.loads(row["data"])
            q["topicId"] = topic_id
            cur.execute(
                "UPDATE questions SET topic_id = ?, data = ? WHERE id = ?",
                (topic_id, json.dumps(q), row["id"]),
            )
        conn.commit()
        conn.close()
        return {"status": "moved", "count": len(rows), "topic_id": topic_id}
    if action == "change_type":
        target_type = data.get("target_type")
        if not target_type:
            conn.close()
            raise HTTPException(status_code=400, detail="Missing target_type")
        placeholders = ",".join(["?"] * len(ids))
        cur.execute(f"SELECT id, data FROM questions WHERE id IN ({placeholders})", ids)
        rows = [dict(row) for row in cur.fetchall()]
        updated = 0
        for row in rows:
            q = json.loads(row["data"])
            q = convert_question_type(q, target_type)
            cur.execute(
                "UPDATE questions SET type = ?, data = ? WHERE id = ?",
                (q["type"], json.dumps(q), row["id"]),
            )
            updated += 1
        conn.commit()
        conn.close()
        return {"status": "updated", "count": updated, "target_type": target_type}
    conn.close()
    raise HTTPException(status_code=400, detail="Unknown action")


@app.get("/attempts")
def get_attempts(user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM attempts WHERE username = ?", (user,))
    rows = [dict(row) for row in cur.fetchall()]
    conn.close()
    return {"attempts": rows}


@app.post("/attempts")
def create_attempt(data: dict, user=Depends(get_current_user)):
    if "questionId" not in data:
        raise HTTPException(status_code=400, detail="Missing questionId")
    attempt_id = data.get("id") or f"att_{int(datetime.now().timestamp()*1000)}"
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO attempts (id, username, question_id, selected, correct, graded_by_user, timestamp, time_ms)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            attempt_id,
            user,
            data["questionId"],
            json.dumps(data.get("selected")),
            1 if data.get("correct") else 0,
            1 if data.get("graded_by_user") else 0,
            data.get("timestamp") or datetime.now(timezone.utc).isoformat(),
            data.get("timeMs"),
        ),
    )
    conn.commit()
    conn.close()
    return {"status": "created", "id": attempt_id}


@app.post("/questions/{question_id}/grade")
def grade_question_endpoint(question_id: str, data: dict, user=Depends(get_current_user)):
    conn = get_db()
    question = get_question_record(conn, question_id)
    conn.close()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    answer = data.get("answer") or {}
    can_auto, correct, expected, debug = grade_question(question, answer)
    if not can_auto:
        return {
            "status": "needs_self_grade",
            "correct": None,
            "expected": expected,
            "feedback": "Self grading required",
            "solution": question.get("solution"),
            "normalized_answer": answer,
        }
    return {
        "status": "graded",
        "correct": bool(correct),
        "expected": expected,
        "feedback": "",
        "solution": question.get("solution"),
        "normalized_answer": answer,
        "debug": debug,
    }


@app.post("/questions/{question_id}/instantiate")
def instantiate_question(question_id: str, data: dict, user=Depends(get_current_user)):
    seed = data.get("seed")
    variant_id = data.get("variant_id")
    mode = data.get("mode", "variant")
    if seed is None:
        seed = int(time.time() * 1000) % 100000
    conn = get_db()
    question = get_question_record(conn, question_id)
    conn.close()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")

    rng = random.Random(seed)
    resolved = json.loads(json.dumps(question))
    params = {}

    variants = resolved.get("variants") or []
    if variants:
        chosen = None
        if variant_id:
            for v in variants:
                if v.get("variant_id") == variant_id:
                    chosen = v
                    break
        if not chosen:
            idx = seed % len(variants)
            chosen = variants[idx]
        overrides = chosen.get("param_overrides") or {}
        expected = chosen.get("expected") or {}
        params.update(overrides)
        if overrides:
            resolved.setdefault("payload", {}).update(overrides)
        if expected:
            resolved.setdefault("solution", {}).update(expected)

    randomization = resolved.get("randomization") or {}
    if randomization:
        ranges = randomization.get("ranges") or {}
        for key, spec in ranges.items():
            min_v = spec.get("min")
            max_v = spec.get("max")
            if isinstance(min_v, (int, float)) and isinstance(max_v, (int, float)):
                params[key] = rng.uniform(min_v, max_v)
        # Replace placeholders in prompt/payload/solution.
        if params:
            def replace_placeholders(text):
                for k, v in params.items():
                    text = text.replace(f"{{{{{k}}}}}", str(v))
                return text
            if isinstance(resolved.get("prompt"), str):
                resolved["prompt"] = replace_placeholders(resolved["prompt"])
            for section in ("payload", "solution", "support"):
                obj = resolved.get(section)
                if isinstance(obj, dict):
                    resolved[section] = json.loads(
                        replace_placeholders(json.dumps(obj))
                    )

    return {
        "instance_id": f"inst_{uuid.uuid4().hex}",
        "seed": seed,
        "question": resolved,
        "params": params,
    }


@app.get("/progress/summary")
def get_progress_summary(user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, topic_id, type FROM questions")
    questions = [dict(row) for row in cur.fetchall()]
    cur.execute("SELECT * FROM attempts WHERE username = ?", (user,))
    attempts_rows = [dict(row) for row in cur.fetchall()]
    conn.close()

    attempts_by_q = {}
    for att in attempts_rows:
        attempts_by_q.setdefault(att["question_id"], []).append(att)
    for att_list in attempts_by_q.values():
        att_list.sort(key=lambda a: a.get("timestamp") or "")

    totals = {
        "questions_total": len(questions),
        "attempted_unique": len(attempts_by_q),
        "correct_unique": len([k for k, v in attempts_by_q.items() if any(a["correct"] for a in v)]),
    }

    by_topic = {}
    by_type = {}

    for q in questions:
        qid = q["id"]
        topic = q["topic_id"] or "unknown"
        qtype = q["type"] or "unknown"
        atts = attempts_by_q.get(qid, [])
        attempted = len(atts) > 0
        correct = sum(1 for a in atts if a["correct"])
        accuracy = correct / len(atts) if atts else 0
        last_three = atts[-3:] if len(atts) >= 3 else atts
        mastery = len(last_three) == 3 and all(a["correct"] for a in last_three)

        for bucket, key in ((by_topic, topic), (by_type, qtype)):
            entry = bucket.setdefault(
                key,
                {
                    "attempted": 0,
                    "correct": 0,
                    "accuracy": 0,
                    "mastery_count": 0,
                    "questions_total": 0,
                },
            )
            entry["questions_total"] += 1
            if attempted:
                entry["attempted"] += 1
            entry["correct"] += correct
            entry["mastery_count"] += 1 if mastery else 0

    for bucket in (by_topic, by_type):
        for entry in bucket.values():
            entry["accuracy"] = (
                entry["correct"] / entry["attempted"] if entry["attempted"] else 0
            )

    return {"totals": totals, "by_topic": by_topic, "by_type": by_type}


@app.get("/appointments")
def get_appointments(user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM appointments WHERE username = ?", (user,))
    rows = [dict(row) for row in cur.fetchall()]
    conn.close()
    return {"appointments": rows}


@app.post("/appointments")
def create_appointment(data: dict, user=Depends(get_current_user)):
    if "title" not in data or "start" not in data:
        raise HTTPException(status_code=400, detail="Missing fields")
    appt_id = data.get("id") or f"appt_{int(datetime.now().timestamp()*1000)}"
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO appointments (id, username, title, start, end, notes, category)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            appt_id,
            user,
            data["title"],
            data["start"],
            data.get("end", ""),
            data.get("notes", ""),
            data.get("category", ""),
        ),
    )
    conn.commit()
    conn.close()
    return {"status": "created", "id": appt_id}


@app.put("/appointments/{appt_id}")
def update_appointment(appt_id: str, data: dict, user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE appointments SET title = ?, start = ?, end = ?, notes = ?, category = ?
        WHERE id = ? AND username = ?
        """,
        (
            data.get("title"),
            data.get("start"),
            data.get("end", ""),
            data.get("notes", ""),
            data.get("category", ""),
            appt_id,
            user,
        ),
    )
    conn.commit()
    conn.close()
    return {"status": "updated"}


@app.delete("/appointments/{appt_id}")
def delete_appointment(appt_id: str, user=Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM appointments WHERE id = ? AND username = ?",
        (appt_id, user),
    )
    conn.commit()
    conn.close()
    return {"status": "deleted"}


@app.post("/ai/ingest")
def ai_ingest(file: UploadFile = File(...), user=Depends(get_current_user)):
    text, content_type, sha256 = extract_text_from_upload(file)
    if not text or len(text) < 20:
        raise HTTPException(status_code=400, detail="Empty document")
    doc_id = f"doc_{uuid.uuid4().hex}"
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO documents (id, filename, sha256, content_type, uploaded_at) VALUES (?, ?, ?, ?, ?)",
        (doc_id, file.filename, sha256, content_type, datetime.now(timezone.utc).isoformat()),
    )
    chunks = chunk_text(text)
    for idx, chunk in enumerate(chunks):
        chunk_id = f"chunk_{uuid.uuid4().hex}"
        emb = ollama_embed(chunk)
        cur.execute(
            "INSERT INTO chunks (id, doc_id, chunk_index, text, embedding) VALUES (?, ?, ?, ?, ?)",
            (chunk_id, doc_id, idx, chunk, json.dumps(emb)),
        )
    conn.commit()
    conn.close()
    log_ai(user, "ingest", {"doc_id": doc_id, "chunks": len(chunks)})
    return {"status": "ok", "doc_id": doc_id, "chunks": len(chunks)}


@app.post("/ai/ask")
def ai_ask(data: dict, user=Depends(get_current_user)):
    query = data.get("query", "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="Missing query")
    top_k = int(data.get("top_k", 5))
    retrieved = retrieve_chunks(query, top_k=top_k)
    context = "\n\n".join(
        [f"[{r['doc_id']}::{r['chunk_id']}] {r['text']}" for r in retrieved]
    )
    system = (
        "You are a study assistant. Use ONLY the provided context. "
        "If context is insufficient, say so. Keep answers concise."
    )
    prompt = f"Question: {query}\n\nContext:\n{context}\n\nAnswer:"
    answer = ollama_generate(prompt, AI_MODELS["chat"], system=system)
    log_ai(user, "ask", {"query": query})
    return {
        "answer": answer,
        "citations": [{"doc_id": r["doc_id"], "chunk_id": r["chunk_id"]} for r in retrieved],
    }


@app.post("/ai/summarize")
def ai_summarize(data: dict, user=Depends(get_current_user)):
    text = data.get("text")
    doc_id = data.get("doc_id")
    if not text and not doc_id:
        raise HTTPException(status_code=400, detail="Provide text or doc_id")
    if doc_id:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT text FROM chunks WHERE doc_id = ? ORDER BY chunk_index", (doc_id,))
        text = "\n".join([row["text"] for row in cur.fetchall()][:6])
        conn.close()
    system = "Summarize clearly in bullet points."
    prompt = f"Summarize:\n{text}"
    summary = ollama_generate(prompt, AI_MODELS["chat"], system=system)
    log_ai(user, "summarize", {"doc_id": doc_id})
    return {"summary": summary}


@app.post("/ai/generate_quiz")
def ai_generate_quiz(data: dict, user=Depends(get_current_user)):
    topic = data.get("topic", "").strip()
    difficulty = data.get("difficulty", "medium")
    count = int(data.get("count", 5))
    store = bool(data.get("store", False))
    if not topic:
        raise HTTPException(status_code=400, detail="Missing topic")
    system = "You output strict JSON only. No commentary."
    prompt = (
        f"Generate {count} quiz questions about {topic} at {difficulty} difficulty.\n"
        "Use JSON array with objects: "
        '{"id":"q_x","type":"single|multi|truefalse|fillblank|explain",'
        '"prompt":"...",'
        '"options":["A","B","C","D"] (only for single/multi),'
        '"correct":[0] (for single/multi),'
        '"correct":true/false (for truefalse),'
        '"answers":["..."] (for fillblank),'
        '"expectedAnswer":"..." (for explain),'
        '"explanation":"..."}'
    )
    raw = ollama_generate(prompt, AI_MODELS["quiz"], system=system, temperature=0.3)
    try:
        questions = json.loads(raw)
        if not isinstance(questions, list):
            raise ValueError("Not a list")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Quiz parse failed: {exc}")
    if store:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM topics WHERE id = ?", (topic,))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO topics (id, name, topic_area, parent_id, path, depth) VALUES (?, ?, ?, ?, ?, ?)",
                (topic, topic.replace("-", " ").title(), "", None, topic, 0),
            )
        for q in questions:
            qid = q.get("id") or f"ai_{uuid.uuid4().hex[:8]}"
            q["id"] = qid
            q["topicId"] = topic
            cur.execute(
                "INSERT OR REPLACE INTO questions (id, topic_id, type, data, explanation, source_ref, tags) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    qid,
                    topic,
                    q.get("type", "single"),
                    json.dumps(q),
                    q.get("explanation", ""),
                    "ai:generated",
                    json.dumps(q.get("tags", [])),
                ),
            )
        conn.commit()
        conn.close()
    log_ai(user, "generate_quiz", {"topic": topic, "count": count})
    return {"questions": questions, "stored": store}


def grade_deterministic(question, answer):
    q_type = question.get("type")
    if q_type == "single":
        return int(answer) == int(question.get("correctIndex"))
    if q_type == "multi":
        correct = set(question.get("correctIndexes", []))
        given = set(answer if isinstance(answer, list) else [])
        return correct == given
    if q_type == "truefalse":
        return bool(answer) == bool(question.get("correctBoolean"))
    if q_type == "fillblank":
        expected = [str(a).strip().lower() for a in question.get("expectedAnswers", [])]
        return str(answer).strip().lower() in expected
    if q_type == "guess":
        expected = [str(a).strip().lower() for a in question.get("expectedAnswers", [])]
        return str(answer).strip().lower() in expected
    if q_type == "ordering":
        return answer == question.get("correct_order")
    return None


@app.post("/ai/grade_answer")
def ai_grade_answer(data: dict, user=Depends(get_current_user)):
    question_id = data.get("question_id")
    answer = data.get("answer")
    if not question_id:
        raise HTTPException(status_code=400, detail="Missing question_id")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT data FROM questions WHERE id = ?", (question_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Question not found")
    question = json.loads(row["data"])
    deterministic = grade_deterministic(question, answer)
    if deterministic is not None:
        return {"correct": deterministic, "confidence": 1.0, "needs_review": False}
    system = "You are a strict grader. Return JSON with keys: correct (true/false), confidence (0-1), needs_review (true/false)."
    prompt = f"Question: {question}\nAnswer: {answer}\nGrade now."
    raw = ollama_generate(prompt, AI_MODELS["grade"], system=system, temperature=0.0)
    try:
        graded = json.loads(raw)
    except Exception:
        graded = {"correct": False, "confidence": 0.0, "needs_review": True}
    return graded


@app.post("/ai/browse")
def ai_browse(data: dict, user=Depends(get_current_user)):
    url = data.get("url")
    query = data.get("query", "")
    if not url:
        raise HTTPException(status_code=400, detail="Missing url")
    check_rate_limit(user, "browse")
    text = fetch_url_text(url)
    system = "Summarize or answer using the fetched page only. Be concise."
    prompt = f"Query: {query}\n\nPage text:\n{text[:6000]}"
    summary = ollama_generate(prompt, AI_MODELS["chat"], system=system, temperature=0.2)
    log_ai(user, "browse", {"url": url})
    return {"summary": summary, "source_url": url}
